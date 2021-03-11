import openpyxl
i=1
#change path to your local directory
path = 'c:/Users/Andrew A/Desktop/python_demo/7_Working_with_Excel/'
book = openpyxl.load_workbook(path + "revenue.xlsx")
tab1 = book.get_sheet_by_name("northamerica")
while str(tab1.cell(row=i,column=1).value)!="None":
    print(str(tab1.cell(row=i,column=1).value)+", "+str(tab1.cell(row=i,column=2).value))
    i=i+1