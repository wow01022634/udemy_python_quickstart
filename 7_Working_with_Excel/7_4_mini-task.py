import openpyxl 
import requests

def url_status(url):  # checks status for URL
    r = requests.get(url)
    return  str(r.status_code)

#load the same excel file
#change path to your local directory
path = 'c:/Users/Andrew A/Desktop/python_demo/7_Working_with_Excel/'
book = openpyxl.load_workbook(path+"404checker.xlsx")
#create a new tab (sheet) named 'April'
tab1=book.get_sheet_by_name("url")
i=2 #define i for iteration as firt row is header
status="" #define x for 404 status
webaddress = "" #define web address
rows = [] #define list to collect list of URLs and 404 status
#create a new tab naned 'URL_status'
sheet = book.create_sheet("URL_status")
while i<=4:
    webaddress = tab1.cell(row=i,column=1).value
    if str(webaddress) != "None":
        sheet.cell(row=i,column=1).value = webaddress
        sheet.cell(row=i,column=2).value = url_status(webaddress)
    i=i+1
#now save it back
book.save(path + "404checker.xlsx")
   